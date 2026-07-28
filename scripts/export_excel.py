import json,sys
from datetime import date,datetime,timezone
from pathlib import Path
from openpyxl import load_workbook
def iso(v):return v.isoformat() if isinstance(v,(datetime,date)) else None
source,target=Path(sys.argv[1]).resolve(),Path(sys.argv[2]).resolve();w=load_workbook(source,read_only=True,data_only=True)
base=[{"id":str(r[0]),"provision":str(r[1]).strip(),"annual":float(r[2] or 0),"unit":str(r[3] or "").strip()} for r in w["01_BASE_TECNICA_PROVISIONES"].iter_rows(min_row=2,values_only=True) if r[0] and r[1]]
records=[]
for r in w["00_REGISTRO_PROVISIONES"].iter_rows(min_row=2,values_only=True):
 if r[0] and r[8] and r[1] in (2025,2026):records.append({"id":str(r[0]),"contractYear":int(r[1]),"requestDate":iso(r[2]),"dueDate":iso(r[3]),"provision":str(r[8]).strip(),"category":str(r[9] or r[8]).strip(),"quantity":float(r[10] or 0),"unit":str(r[11] or "").strip(),"status":str(r[12] or "Sin estado").strip(),"observations":str(r[14] or "").strip()})
target.parent.mkdir(parents=True,exist_ok=True);target.write_text(json.dumps({"generatedAt":datetime.now(timezone.utc).isoformat(),"base":base,"records":records},ensure_ascii=False,indent=2),encoding="utf-8")
print(f"{len(base)} provisiones base y {len(records)} movimientos")
